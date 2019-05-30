import os
import win32com.client
import random
from openpyxl import *
from openpyxl.utils.cell import get_column_letter, column_index_from_string

##########get excel index############
def get_excel_indices(ws, index_headings_col, index_values_col, index_start_row):
    excel_index = {}
    current_row = index_start_row
    while ws[index_headings_col + str(current_row)].value is not None:
        index_heading = ws[index_headings_col + str(current_row)].value
        index_value = ws[index_values_col + str(current_row)].value
        #enter the new entry into the index
        excel_index[index_heading] = index_value
        current_row = current_row + 1
    return excel_index

wb = load_workbook('SetupAB.xlsx')
ws_index = wb.get_sheet_by_name('Index')
ExcelIndex = get_excel_indices(ws_index, 'A', 'B', 2)

InputTable = ExcelIndex['Input table sheet']
FloorPlan = ExcelIndex['Floor plans sheet']
SectionProperties = ExcelIndex['Section properties sheet']
Bracing = ExcelIndex['Bracing sheet']
FloorBracing = ExcelIndex['Floor bracing sheet']
Materials = ExcelIndex['Materials sheet']
InputTableOffset = ExcelIndex['Input table offset']
PropertiesStartRow = ExcelIndex['Properties start row']

for keys,values in ExcelIndex.items():
    print(keys)
    print(values)

##########read input table############
def get_section_properties(ws,section_headings_start_col, section_values_start_col, section_start_row):
    
    #while cell is not none
        #column_index_from_string() + 3
        #get_column_letter()
    section_type = {};
    current_property_col = section_headings_start_col;
    current_value_col = section_values_start_col;
    i = 1 
    while ws[current_property_col+str(1)].value is not None:
        current_row = section_start_row
        section_type['Section'+str(i)]={}
        while ws[current_property_col + str(current_row)].value is not None:
            section_properties = {}
            section_properties_heading = ws[current_property_col + str(current_row)].value
            section_properties_value = ws[current_value_col + str(current_row)].value

            #enter the new entry into the index
            section_type['Section'+str(i)][section_properties_heading] = section_properties_value
            current_row = current_row + 1
        i += 1
        current_property_col = get_column_letter(column_index_from_string(current_property_col)+3)
        current_value_col = get_column_letter(column_index_from_string(current_value_col)+3)
    return section_type

ws_section = wb.get_sheet_by_name('Section Properties')
SectionProperties = get_section_properties(ws_section, 'A', 'B',4) #or use ExcelIndex 

for keys,values in SectionProperties.items():
    print(keys)
    print(values)

#def read_input_table(ws,floor_col,properties_col,floor_row):


