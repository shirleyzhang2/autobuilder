import os
import win32com.client
import random
from openpyxl import *
from openpyxl.utils.cell import get_column_letter, column_index_from_string
import string

class Tower:
    def __init__(self, number, member_mat, rod_mat, footprint, floor_plans, floor_heights, col_props, bracing_types, floor_masses, floor_bracing_types,side):
        self.number = number
        self.member_mat = member_mat
        self.rod_mat = rod_mat
        self.footprint = footprint
        self.floor_plans = floor_plans
        self.floor_heights = floor_heights
        self.col_props = col_props
        self.bracing_types = bracing_types
        self.floor_masses = floor_masses
        self.floor_bracing_types = floor_bracing_types
        self.side = side

class BracingScheme:
    def __init__(self, number=1, face=1, members=[]):
        self.number = number
        self.members = members

class FloorPlan:
    def __init__(self, number=1, members=[], mass_nodes=[], scaling_x = 0, scaling_y = 0, area = 0):
        self.number = number
        self.members = members
        self.mass_nodes = mass_nodes
        self.scaling_x = scaling_x
        self.scaling_y = scaling_y
        self.area = area

class Member:
    def __init__(self, start_node=[], end_node=[], sec_prop=1):
        self.start_node = start_node
        self.end_node = end_node
        self.sec_prop = sec_prop


##########get excel index############
def get_excel_indices(wb, index_headings_col, index_values_col, index_start_row):
    excel_index = {}
    ws = wb['Index']
    current_row = index_start_row
    while ws[index_headings_col + str(current_row)].value is not None:
        index_heading = ws[index_headings_col + str(current_row)].value
        index_value = ws[index_values_col + str(current_row)].value
        #enter the new entry into the index
        excel_index[index_heading] = index_value
        current_row = current_row + 1
    return excel_index

##########read excel tabs############
def get_properties(wb,excel_index,parameter):
    headings_start_col = excel_index['Section or material properties col']
    values_start_col = excel_index['Section or material values col']
    start_row = excel_index['Properties start row']
    if parameter == 'Material':
        ws = wb['Materials']
    elif parameter == 'Section':
        ws = wb['Section Properties']
    else:
        print('Input should be either "Material" or"Section"')
    #parameter = 'unknown'
    #if ws['A1'].value == 'Section #':
    #    parameter = 'Section'
    #else:
    #    parameter = 'Material'
    parameter_type = {};
    current_property_col = headings_start_col;
    current_value_col = values_start_col;
    i = 1 
    while ws[current_property_col+str(1)].value is not None:
        current_row = start_row
        parameter_type[parameter+' '+str(i)]={}
        while ws[current_property_col + str(current_row)].value is not None:
            properties = {}
            properties_heading = ws[current_property_col + str(current_row)].value
            properties_value = ws[current_value_col + str(current_row)].value
            #enter the new entry into the index
            parameter_type[parameter+' '+str(i)][properties_heading] = properties_value
            current_row = current_row + 1
        i += 1
        current_property_col = get_column_letter(column_index_from_string(current_property_col)+3)
        current_value_col = get_column_letter(column_index_from_string(current_value_col)+3)
    return parameter_type

#Outputs a list containing tower objects representing each tower to be built
def read_input_table(wb,excel_index):
    #Read in the top value in the sheet
    input_table_sheet = excel_index['Input table sheet']
    #Set worksheet to the input table sheet
    ws_input = wb['Input Table']
    input_table_offset = excel_index['Input table offset']
    total_towers = excel_index['Total number of towers']
    cur_tower_row = 1
    table_start_row = cur_tower_row + input_table_offset
    all_towers = []
    cur_tower_num = 1
    while cur_tower_num <= total_towers:
        #create tower object
        cur_tower = Tower(0,0,0,0,0,0,0,0,0,0,0)
        #read member material
        member_mat = ws_input['B'+str(cur_tower_row + 1)].value
        cur_tower.member_mat = member_mat
        #read rod material
        rod_mat = ws_input['B'+str(cur_tower_row + 2)].value
        cur_tower.rod_mat = rod_mat
        #read building footprint in square inches
        footprint = ws_input['B'+str(cur_tower_row + 3)].value
        cur_tower.footprint = footprint
        #read floor plans
        floor_pln_col = excel_index['Floor plan col']
        cur_floor_row = cur_tower_row + input_table_offset
        floor_plans = []
        while ws_input[floor_pln_col + str(cur_floor_row)].value is not None:
            floor_pln = ws_input[floor_pln_col + str(cur_floor_row)].value
            floor_plans.append(floor_pln)
            cur_floor_row = cur_floor_row + 1
        cur_tower.floor_plans = floor_plans
        #read floor heights
        floor_heights_col = excel_index['Floor height col']
        cur_floor_row = cur_tower_row + input_table_offset
        floor_heights = []
        while ws_input[floor_heights_col + str(cur_floor_row)].value is not None:
            floor_height = ws_input[floor_heights_col + str(cur_floor_row)].value
            floor_heights.append(floor_height)
            cur_floor_row = cur_floor_row + 1
        cur_tower.floor_heights = floor_heights
        #read column properties
        col_props_start_col = excel_index['Column properties start']
        col_props_end_col = excel_index['Column properties end']
        cur_floor_row = cur_tower_row + input_table_offset
        cur_col = col_props_start_col
        col_props = {}
        face = 1
        while ws_input[cur_col + str(cur_floor_row)].value is not None:
            col_props[face] = []
            while cur_col != get_column_letter(column_index_from_string(col_props_end_col)+1):
                col_prop = ws_input[cur_col + str(cur_floor_row)].value
                col_props[face] = col_prop
                cur_col = get_column_letter(column_index_from_string(cur_col)+1)
            cur_col = col_props_start_col
            cur_floor_row = cur_floor_row + 1
        cur_tower.col_props = col_props
        #read bracing properties
        bracing_types_start_col = excel_index['Bracing type start']
        bracing_types_end_col= excel_index['Bracing type end']
        cur_floor_row = cur_tower_row + input_table_offset
        cur_col = bracing_types_start_col
        bracing_types = {}
        face = 1
        while ws_input[cur_col + str(cur_floor_row)].value is not None:
            bracing_types[face] = []
            while cur_col != get_column_letter(column_index_from_string(bracing_types_end_col)+1):
                bracing_type = ws_input[cur_col + str(cur_floor_row)].value
                bracing_types[face].append(bracing_type)
                cur_col = get_column_letter(column_index_from_string(cur_col)+1)
            cur_col = bracing_types_start_col
            cur_floor_row = cur_floor_row + 1
            face += 1
        cur_tower.bracing_types = bracing_types
        #read floor masses
        floor_masses_col = excel_index['Floor mass col']
        cur_floor_row = cur_tower_row + input_table_offset
        floor_masses = []
        while ws_input[floor_masses_col + str(cur_floor_row)].value is not None:
            floor_mass = ws_input[floor_masses_col + str(cur_floor_row)].value
            floor_masses.append(floor_mass)
            cur_floor_row = cur_floor_row + 1
        cur_tower.floor_masses = floor_masses
        #read floor bracing types
        floor_bracing_col = excel_index['Floor bracing col']
        cur_floor_row = cur_tower_row + input_table_offset
        floor_bracing_types = []
        while ws_input[floor_bracing_col + str(cur_floor_row)].value is not None:
            floor_bracing = ws_input[floor_bracing_col + str(cur_floor_row)].value
            floor_bracing_types.append(floor_bracing)
            cur_floor_row = cur_floor_row + 1
        cur_tower.floor_bracing_types = floor_bracing_types
        #read number of sides
        side = []
        side_start_col = excel_index['Bracing type start']
        side_end_col= excel_index['Bracing type end']
        cur_side_row = cur_tower_row + input_table_offset - 1
        cur_col = side_start_col
        while cur_col != get_column_letter(column_index_from_string(side_end_col)+1):
            side_num = ws_input[cur_col + str(cur_side_row)].value
            side.append(side_num)
            cur_col = get_column_letter(column_index_from_string(cur_col)+1)
        cur_tower.side = side
        #increment
        cur_tower.number = cur_tower_num
        all_towers.append(cur_tower)
        cur_tower_num = cur_tower_num + 1
        cur_tower_row = cur_floor_row + 1
    return all_towers

def get_node_info(wb,excel_index,node_num_col,parameter):
    if parameter == 'Floor Bracing':
        ws = wb['Floor Bracing']
    elif parameter == 'Bracing':
        ws = wb['Bracing']
    elif parameter == 'Floor Plans':
        ws = wb['Floor Plans']
    horiz_col = get_column_letter(column_index_from_string(node_num_col)+1)
    vert_col = get_column_letter(column_index_from_string(node_num_col)+2)
    mass_col = get_column_letter(column_index_from_string(node_num_col)+3)
    start_row = excel_index['Properties start row']
    nodes = []
    mass_nodes = []
    current_row = start_row
    while ws[node_num_col + str(current_row)].value is not None:
        horiz = ws[horiz_col + str(current_row)].value
        vert = ws[vert_col + str(current_row)].value
        if parameter == 'Floor Plans':
            mass_at_node = ws[mass_col + str(current_row)].value
            if mass_at_node ==1:
                mass_nodes.append([horiz,vert])
        #enter the new entry into the list
        nodes.append([horiz, vert])
        current_row = current_row + 1
    return nodes, mass_nodes

def get_bracing(wb,excel_index,parameter):
    headings_col = excel_index['Bracing start col']
    section_col = excel_index['Bracing section col']
    start_node_col = excel_index['Bracing start node col']
    end_node_col = excel_index['Bracing end node col']
    start_row = excel_index['Properties start row']
    if parameter == 'Floor Bracing':
        ws = wb['Floor Bracing']
    elif parameter == 'Bracing':
        ws = wb['Bracing']
    else:
        print('Input should be either "Floor Bracing" or "Bracing"')
    all_bracing = []
    current_headings_col = headings_col
    current_section_col = section_col
    current_start_node_col = start_node_col
    current_end_node_col = end_node_col
    i = 1
    while ws[current_headings_col+str(4)].value is not None:
        nodes = get_node_info(wb, excel_index, current_headings_col, parameter)[0]
        current_row = start_row
        j = 1
        cur_members = []
        while ws[current_start_node_col + str(current_row)].value is not None:
            section = ws[current_section_col + str(current_row)].value
            start_node_num = ws[current_start_node_col + str(current_row)].value
            end_node_num = ws[current_end_node_col + str(current_row)].value
            start_node = nodes[start_node_num-1]
            end_node = nodes[end_node_num-1]
            cur_members.append(Member(start_node, end_node, section))
            current_row = current_row + 1
            j += 1
        all_bracing.append(BracingScheme(number=i, members=cur_members))
        i += 1
        current_headings_col = get_column_letter(column_index_from_string(current_headings_col)+8)
        current_section_col = get_column_letter(column_index_from_string(current_section_col)+8)
        current_start_node_col = get_column_letter(column_index_from_string(current_start_node_col)+8)
        current_end_node_col = get_column_letter(column_index_from_string(current_end_node_col)+8)
    return all_bracing

def get_floor_plans(wb,excel_index):
    headings_col = excel_index['Floor plan start col']
    section_col = excel_index['Floor plan section col']
    start_node_col = excel_index['Floor plan start node col']
    end_node_col = excel_index['Floor plan end node col']
    start_row = excel_index['Properties start row']
    ws = wb['Floor Plans']
    all_plans = []
    current_headings_col = headings_col
    current_section_col = section_col
    current_start_node_col = start_node_col
    current_end_node_col = end_node_col
    i = 1
    while ws[current_headings_col + str(4)].value is not None:
        [nodes, mass_nodes] = get_node_info(wb, excel_index, current_headings_col, 'Floor Plans')
        current_row = start_row

        cur_members = []
        max_node_x = 0
        max_node_y = 0
        min_node_x = 0
        min_node_y = 0
        while ws[current_start_node_col + str(current_row)].value is not None:
            section = ws[current_section_col + str(current_row)].value
            start_node_num = ws[current_start_node_col + str(current_row)].value
            end_node_num = ws[current_end_node_col + str(current_row)].value
            start_node = nodes[start_node_num - 1]
            end_node = nodes[end_node_num - 1]
            cur_members.append(Member(start_node, end_node, section))
            #find scaling factor in x and y, find area
            cur_nodes = []
            cur_nodes.append(start_node)
            cur_nodes.append(end_node)
            for node in cur_nodes:
                if max_node_x < start_node[0]:
                    max_node_x = node[0]
                if max_node_y < node[1]:
                    max_node_y = node[1]
                if min_node_x > node[0]:
                    min_node_x = node[0]
                if min_node_y > node[1]:
                    min_node_y = node[1]
            current_row = current_row + 1

        scaling_x = max_node_x - min_node_x
        scaling_y = max_node_y - min_node_y
        #area = scaling_x*scaling_y
        
        all_plans.append(FloorPlan(number=i, members=cur_members, mass_nodes=mass_nodes, scaling_x = scaling_x, scaling_y = scaling_y))
        i += 1
        current_headings_col = get_column_letter(column_index_from_string(current_headings_col) + 9)
        current_section_col = get_column_letter(column_index_from_string(current_section_col) + 9)
        current_start_node_col = get_column_letter(column_index_from_string(current_start_node_col) + 9)
        current_end_node_col = get_column_letter(column_index_from_string(current_end_node_col) + 9)
    return all_plans


#TESTING
#wb = load_workbook('SetupAB.xlsm')
#ExcelIndex = get_excel_indices(wb, 'A', 'B', 2)

#InputTable = ExcelIndex['Input table sheet']
#FloorPlan = ExcelIndex['Floor plans sheet']
#SectionProperties = ExcelIndex['Section properties sheet']
#Bracing = ExcelIndex['Bracing sheet'
#FloorBracing = ExcelIndex['Floor bracing sheet']
#Materials = ExcelIndex['Materials sheet']
#InputTableOffset = ExcelIndex['Input table offset']
#PropertiesStartRow = ExcelIndex['Properties start row']

#sectionproperties = get_properties(wb,excelindex,'section')
#materials = get_properties(wb,excelindex,'material')
#bracingschemes = get_bracing(wb,excelindex,'bracing')


#AllFloorPlans = get_floor_plans(wb,ExcelIndex)

#FloorBracing = get_floor_or_bracing(wb,ExcelIndex,'Floor Bracing')

#for keys,values in ExcelIndex.items():
#    print(keys)
#    print(values)

#for keys,values in SectionProperties.items():
#    print(keys)
#    print(values)

#for keys,values in Materials.items():
#    print(keys)
#    print(values)

#for keys,values in Bracing.items():
#   print(keys) 
#   print(values)


#for FloorPlan in AllFloorPlans:
#    print('number ' + str(FloorPlan.number))
#    for member in FloorPlan.members:
#        print(member.start_node)
#        print(member.end_node)
#        print(member.sec_prop)
#    print(FloorPlan.mass_nodes)
#    print(FloorPlan.scaling_x)
#    print(FloorPlan.scaling_y)
#    print(FloorPlan.area)

#for Scheme in BracingSchemes:
    #print('number ' + str(Scheme.number))
    #for member in Scheme.members:        
        #print(member.start_node)
        #print(member.end_node)
        #print(member.sec_prop)


#AllTowers = read_input_table(wb, ExcelIndex)
#for tower in AllTowers:
#    print(tower.number)
#    print(tower.member_mat)
#    print(tower.rod_mat)
#    print(tower.floor_plans)
#    print(tower.floor_heights)
#    print(tower.col_props)
#    print(tower.bracing_types)
#    print(tower.floor_masses)
#    print(tower.floor_bracing_types)
