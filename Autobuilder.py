import os
import win32com.client
import openpyxl
import random
from openpyxl import *
import re
import time
import ReadExcel
import scipy
import numpy
from scipy.stats import norm
import datetime
import matplotlib.pyplot as plt


def build_floor_plan_and_bracing(SapModel, tower, all_floor_plans, all_floor_bracing, floor_num, floor_elev):
    print('Building floor plan...')
    floor_plan_num = tower.floor_plans[floor_num-1]
    floor_plan = all_floor_plans[floor_plan_num-1]
    #Create members for floor plan
    for member in floor_plan.members:
        kip_in_F = 3
        SapModel.SetPresentUnits(kip_in_F)
        start_node = member.start_node
        end_node = member.end_node
        start_x = start_node[0]
        start_y = start_node[1]
        start_z = floor_elev
        end_x = end_node[0]
        end_y = end_node[1]
        end_z = floor_elev
        section_name = member.sec_prop
        [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, PropName=section_name)
        if ret != 0:
            print('ERROR creating floor plan member on floor ' + str(floor_num))
    #assign masses to mass nodes and create steel rod
    mass_node_1 = floor_plan.mass_nodes[0]
    mass_node_2 = floor_plan.mass_nodes[1]
    floor_mass = tower.floor_masses[floor_num-1]
    mass_per_node = floor_mass/2
    #Create the mass node point
    [ret, mass_name_1] = SapModel.PointObj.AddCartesian(mass_node_1[0],mass_node_1[1],floor_elev,MergeOff=False)
    if ret != 0:
        print('ERROR setting mass nodes on floor ' + str(floor_num))
    [ret, mass_name_2] = SapModel.PointObj.AddCartesian(mass_node_2[0],mass_node_2[1],floor_elev,MergeOff=False)
    if ret != 0:
        print('ERROR setting mass nodes on floor ' + str(floor_num))
    #Assign masses to the mass nodes
    #Shaking in the x direcion!
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    ret = SapModel.PointObj.SetMass(mass_name_1, [mass_per_node,0,0,0,0,0],0,True,False)
    if ret[0] != 0:
        print('ERROR setting mass on floor ' + str(floor_num))
    ret = SapModel.PointObj.SetMass(mass_name_2, [mass_per_node,0,0,0,0,0])
    if ret[0] != 0:
        print('ERROR setting mass on floor ' + str(floor_num))
    #Create steel rod
    kip_in_F = 3
    SapModel.SetPresentUnits(kip_in_F)
    [ret, name1] = SapModel.FrameObj.AddByCoord(mass_node_1[0], mass_node_1[1], floor_elev, mass_node_2[0], mass_node_2[1], floor_elev, PropName='Steel rod')
    if ret !=0:
        print('ERROR creating steel rod on floor ' + str(floor_num))
    #Create floor load forces
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    ret = SapModel.PointObj.SetLoadForce(mass_name_1, 'DEAD', [0, 0, mass_per_node*9.81, 0, 0, 0])
    ret = SapModel.PointObj.SetLoadForce(mass_name_2, 'DEAD', [0, 0, mass_per_node*9.81, 0, 0, 0])
    #create floor bracing
    floor_bracing_num = tower.floor_bracing_types[floor_num-1]
    floor_bracing = all_floor_bracing[floor_bracing_num-1]
    #Find scaling factors
    scaling_x = floor_plan.scaling_x
    scaling_y = floor_plan.scaling_y
    #Create floor bracing
    print('Building floor bracing...')
    for member in floor_bracing.members:
        kip_in_F = 3
        SapModel.SetPresentUnits(kip_in_F)
        start_node = member.start_node
        end_node = member.end_node
        start_x = start_node[0] * scaling_x
        start_y = start_node[1] * scaling_y
        start_z = floor_elev
        end_x = end_node[0] * scaling_x
        end_y = end_node[1] * scaling_y
        end_z = floor_elev
        section_name = member.sec_prop
        [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, PropName=section_name)
        if ret != 0:
            print('ERROR creating floor bracing member on floor ' + str(floor_num))
    return SapModel

def build_face_bracing(SapModel, tower, all_floor_plans, all_face_bracing, floor_num, floor_elev):
    print('Building face bracing...')
    i = 1
    while i <= len(Tower.side):
        face_bracing_num = Tower.bracing_types[floor_num][i-1]
        face_bracing = all_face_bracing[face_bracing_num-1]

        #Find scaling factors
        floor_plan_num = tower.floor_plans[floor_num-1]
        floor_plan = all_floor_plans[floor_plan_num-1]
       
        scaling_x = floor_plan.scaling_x
        scaling_y = floor_plan.scaling_y
        scaling_z = tower.floor_heights[floor_num-1]
        
        for member in face_bracing.members:
            kip_in_F = 3
            SapModel.SetPresentUnits(kip_in_F)
            start_node = member.start_node
            end_node = member.end_node
            
            #Create face bracing for long side
            if i == 1 or i == 3:
                scaling_x_or_y = scaling_x
            #Create face bracing for short side
            elif i == 2 or i == 4:
                scaling_x_or_y = scaling_y

            start_x = start_node[0] * scaling_x_or_y
            start_y = 0
            start_z = start_node[1] * scaling_z + floor_elev
            end_x = end_node[0] * scaling_x_or_y
            end_y = 0
            end_z = end_node[1] * scaling_z + floor_elev
            section_name = member.sec_prop 
            #rotate coordinate system through side 1 - 4
            if i == 1:
                ret = SapModel.CoordSys.SetCoordSys('CSys1', 0, 0, 0, 0, 0, 0)
            elif i == 2:
                ret = SapModel.CoordSys.SetCoordSys('CSys1', scaling_x, 0, 0, 90, 0, 0)
            elif i == 3:
                ret = SapModel.CoordSys.SetCoordSys('CSys1', 0, scaling_y, 0, 0, 0, 0)
            elif i == 4:
                ret = SapModel.CoordSys.SetCoordSys('CSys1', 0, 0, 0, 90, 0, 0)

            [ret, name] = SapModel.FrameObj.AddByCoord(start_x, start_y, start_z, end_x, end_y, end_z, ' ', section_name, ' ', 'CSys1')
            if ret != 0:
                print('ERROR creating floor bracing member on floor ' + str(floor_num))
        i += 1
    return SapModel

def set_base_restraints(SapModel):
    # Set fixed ends on all ground level nodes
    node_num = 1
    [ret, number_nodes, all_node_names] = SapModel.PointObj.GetNameList()
    for node_name in all_node_names:
        [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(node_name, 0, 0, 0)
        if z == 0:
            [ret_set_restraint, ret] = SapModel.PointObj.SetRestraint(node_name, [True, True, True, True, True, True])
    return SapModel

def define_loading(SapModel, time_history_loc, save_loc):
    print('Defining loading...')
    # Define time history function
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.Func.FuncTH.SetFromFile('GM', time_history_loc, 1, 0, 1, 2, True)
    # Set the time history load case
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    SapModel.LoadCases.ModHistLinear.SetCase('GM')
    SapModel.LoadCases.ModHistLinear.SetMotionType('GM', 1)
    SapModel.LoadCases.ModHistLinear.SetLoads('GM', 1, ['Accel'], ['U1'], ['GM'], [1], [1], [0], ['Global'], [0])
    SapModel.LoadCases.ModHistLinear.SetTimeStep('GM', 250, 0.1)
    # Create load combination
    SapModel.RespCombo.Add('DEAD + GM', 0)
    SapModel.RespCombo.SetCaseList('DEAD + GM', 0, 'DEAD', 1)
    SapModel.RespCombo.SetCaseList('DEAD + GM', 0, 'GM', 1)
    # Save the model
    ret = SapModel.File.Save(save_loc)
    if ret != 0:
        print('ERROR saving SAP2000 file')
    return SapModel

#returns the max acceleration in g, max drift (displacement) in mm, and weight in pounds
def run_analysis(SapModel):
    #Run Analysis
    print('Computing...')
    SapModel.Analyze.RunAnalysis()
    print('Finished computing.')
    #Get RELATIVE acceleration from node
    SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    SapModel.Results.Setup.SetComboSelectedForOutput('DEAD + GM', True)
    #set type to envelope
    SapModel.Results.Setup.SetOptionModalHist(1)
    #Get joint acceleration
    #Find a node that is on the top floor
    [ret, number_nodes, all_node_names] = SapModel.PointObj.GetNameList()
    z_max = 0
    z = 0
    for node_name in all_node_names:
        [ret, x, y, z] = SapModel.PointObj.GetCoordCartesian(node_name, 0, 0, 0)
        if z > z_max:
            roof_node_name = node_name
            z_max = z
    #Retrieve max accelerations
    #Set units to metres
    N_m_C = 10
    SapModel.SetPresentUnits(N_m_C)
    g = 9.81
    ret = SapModel.Results.JointAccAbs(roof_node_name, 0)
    max_and_min_acc = ret[7]
    max_pos_acc = max_and_min_acc[0]
    min_neg_acc = max_and_min_acc[1]
    if abs(max_pos_acc) >= abs(min_neg_acc):
        max_acc = abs(max_pos_acc)/g
    elif abs(min_neg_acc) >= abs(max_pos_acc):
        max_acc = abs(min_neg_acc)/g
    else:
        print('Could not find max acceleration')
    #Get joint displacement
    #Set units to millimetres
    N_mm_C = 9
    SapModel.SetPresentUnits(N_mm_C)
    ret = SapModel.Results.JointDispl(roof_node_name, 0)
    max_and_min_disp = ret[7]
    max_pos_disp = max_and_min_disp[0]
    min_neg_disp = max_and_min_disp[1]
    if abs(max_pos_disp) >= abs(min_neg_disp):
        max_drift = abs(max_pos_acc)
    elif abs(min_neg_disp) >= abs(max_pos_disp):
        max_drift = abs(min_neg_disp)
    else:
        print('Could not find max drift')
    #Get weight
    #Get base reactions
    SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    SapModel.Results.Setup.SetCaseSelectedForOutput('DEAD')
    #SapModel.Results.BaseReact(NumberResults, LoadCase, StepType, StepNum, Fx, Fy, Fz, Mx, My, Mz, gx, gy, gz)
    ret = SapModel.Results.BaseReact()
    if ret[0] != 0:
        print('ERROR getting base reaction forces')
    base_react = ret[7][0]
    total_weight = base_react / 9.81
    #convert to lb
    total_weight = total_weight / 0.45359237
    return max_acc, max_drift, total_weight


def get_FABI(max_acc, max_disp, footprint, weight):
    design_life = 100 #years
    construction_cost = 2500000*(weight**2)+6*(10**6)
    land_cost = 35000 * footprint
    annual_building_cost = (land_cost + construction_cost) / design_life
    floor_num = len(Tower.floor_heights)
    if floor_num <= 2:
        annual_revenue = 250 * floor_num
    elif floor_num <= 9:
        annual_revenue = 250 * 2 + 175 * (floor_num - 2)
    elif floor_num <= 15:
        annual_revenue = 250 * 2 + 175 * 7 + 225 * (floor_num - 9)
    else:
        annual_revenue = 250 * 2 + 175 * 7 + 225 * 6 + 275 * (floor_num - 15)
    #annual_revenue = 430300
    equipment_cost = 20000000
    return_period_1 = 50
    return_period_2 = 300
    apeak_1 = max_acc #g's
    xpeak_1 = 100*max_disp/1524 #% roof drift
    structural_damage_1 = scipy.stats.norm(1.5, 0.5).cdf(xpeak_1)
    equipment_damage_1 = scipy.stats.norm(1.75, 0.7).cdf(apeak_1)
    economic_loss_1 = structural_damage_1*construction_cost + equipment_damage_1*equipment_cost
    annual_economic_loss_1 = economic_loss_1/return_period_1
    structural_damage_2 = 0.5
    equipment_damage_2 = 0.5
    economic_loss_2 = structural_damage_2*construction_cost + equipment_damage_2*equipment_cost
    annual_economic_loss_2 = economic_loss_2/return_period_2
    annual_seismic_cost = annual_economic_loss_1 + annual_economic_loss_2
    fabi = annual_revenue - annual_building_cost - annual_seismic_cost
    return fabi

def write_to_excel(wb, all_fabi, save_loc):
    print('Writing all results to Excel...')
    filepath = save_loc + '/Results.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Tower #'
    ws['B1'] = 'FABI'
    fabi_num = 1
    for fabi in all_fabi:
        ws['A' + str(fabi_num + 1)].value = fabi_num
        ws['B' + str(fabi_num + 1)].value = fabi
        fabi_num += 1
    wb.save(filepath)




#----START-----------------------------------------------------START----------------------------------------------------#



print('\n--------------------------------------------------------')
print('Autobuilder by University of Toronto Seismic Design Team')
print('--------------------------------------------------------\n')

#Read in the excel workbook
print("\nReading Excel spreadsheet...")
wb = load_workbook('SetupAB.xlsm')
ExcelIndex = ReadExcel.get_excel_indices(wb, 'A', 'B', 2)

Sections = ReadExcel.get_properties(wb,ExcelIndex,'Section')
Materials = ReadExcel.get_properties(wb,ExcelIndex,'Material')
Bracing = ReadExcel.get_bracing(wb,ExcelIndex,'Bracing')
FloorPlans = ReadExcel.get_floor_plans(wb,ExcelIndex)
FloorBracing = ReadExcel.get_bracing(wb,ExcelIndex,'Floor Bracing')
AllTowers = ReadExcel.read_input_table(wb, ExcelIndex)
SaveLoc = ExcelIndex['Save location']
TimeHistoryLoc = ExcelIndex['Time history location']

print('\nInitializing SAP2000 model...')
# create SAP2000 object
SapObject = win32com.client.Dispatch('SAP2000v15.SapObject')
# start SAP2000
SapObject.ApplicationStart()
# create SapModel Object
SapModel = SapObject.SapModel
# initialize model
SapModel.InitializeNewModel()
# create new blank model
ret = SapModel.File.NewBlank()

#Define new materials
print("\nDefining materials...")
N_m_C = 10
SapModel.SetPresentUnits(N_m_C)
for Material, MatProps in Materials.items():
    MatName = MatProps['Name']
    MatType = MatProps['Material type']
    MatWeight = MatProps['Weight per volume']
    MatE = MatProps['Elastic modulus']
    MatPois = MatProps['Poisson\'s ratio']
    MatTherm = MatProps['Thermal coefficient']
    #Create material type
    ret = SapModel.PropMaterial.SetMaterial(MatName, MatType)
    if ret != 0:
        print('ERROR creating material type')
    #Set isotropic material proprties
    ret = SapModel.PropMaterial.SetMPIsotropic(MatName, MatE, MatPois, MatTherm)
    if ret != 0:
        print('ERROR setting material properties')
    #Set unit weight
    ret = SapModel.PropMaterial.SetWeightAndMass(MatName, 1, MatWeight)
    if ret != 0:
        print('ERROR setting material unit weight')

#Define new sections
print('Defining sections...')
kip_in_F = 3
SapModel.SetPresentUnits(kip_in_F)
for Section, SecProps in Sections.items():
    SecName = SecProps['Name']
    SecArea = SecProps['Area']
    SecTors = SecProps['Torsional constant']
    SecIn3 = SecProps['Moment of inertia about 3 axis']
    SecIn2 = SecProps['Moment of inertia about 2 axis']
    SecSh2 = SecProps['Shear area in 2 direction']
    SecSh3 = SecProps['Shear area in 3 direction']
    SecMod3 = SecProps['Section modulus about 3 axis']
    SecMod2 = SecProps['Section modulus about 2 axis']
    SecPlMod3 = SecProps['Plastic modulus about 3 axis']
    SecPlMod2 = SecProps['Plastic modulus about 2 axis']
    SecRadGy3 = SecProps['Radius of gyration about 3 axis']
    SecRadGy2 = SecProps['Radius of gyration about 2 axis']
    SecMat = SecProps['Material']
    #Create section property
    ret = SapModel.PropFrame.SetGeneral(SecName, SecMat, 0.1, 0.1, SecArea, SecSh2, SecSh3, SecTors, SecIn2, SecIn3, SecMod2, SecMod3, SecPlMod2, SecPlMod3, SecRadGy2, SecRadGy3, -1)
    if ret != 0:
        print('ERROR creating section property ' + SecName)

AllFABI = []
TowerNum = 1
ComputeTimes = []

# Define load cases
SapModel = define_loading(SapModel, TimeHistoryLoc, SaveLoc)
# Start scatter plot of FABI
xdata = []
ydata = []
axes = plt.gca()
axes.set_xlim(1, len(AllTowers))
axes.set_ylim(bottom=0)
ScatterPlot, = axes.plot(xdata, ydata, 'ro')
plt.grid(True, 'both', 'both')
plt.xlabel('Tower Number')
plt.ylabel('FABI')
plt.show(block=False)
plt.ion()


StartTime = time.time()
# Build all towers defined in spreadsheet
for Tower in AllTowers:
    print('\nBuilding tower number ' + str(TowerNum))
    print('-------------------------')
    NumFloors = len(Tower.floor_plans)
    CurFloorNum = 1
    CurFloorElevation = 0
    # Build each floor of the tower

    while CurFloorNum <=  NumFloors:
        print('Floor ' + str(CurFloorNum))
        if CurFloorNum <=  NumFloors:
            SapModel = build_floor_plan_and_bracing(SapModel, Tower, FloorPlans, FloorBracing, CurFloorNum, CurFloorElevation)
        if CurFloorNum <  NumFloors:
            SapModel = build_face_bracing(SapModel, Tower, FloorPlans, Bracing, CurFloorNum, CurFloorElevation)
        #INSERT FUNCTION TO CREATE COLUMNS AT CURRENT FLOOR

        CurFloorHeight = Tower.floor_heights[CurFloorNum - 1]
        CurFloorElevation = CurFloorElevation + CurFloorHeight
        CurFloorNum += 1
    # Set fixed end conditions on all ground floor nodes
    SapModel = set_base_restraints(SapModel)
    # Save the file
    SapModel.File.Save(SaveLoc + '/Tower ' + str(TowerNum))
    #Analyse tower and print results to spreadsheet
    print('\nAnalyzing tower number ' + str(TowerNum))
    print('-------------------------')
    #run analysis and get weight and acceleration
    [MaxAcc, MaxDisp, Weight] = run_analysis(SapModel)
    #Calculate model FABI
    AllFABI.append(get_FABI(MaxAcc, MaxDisp, Tower.footprint, Weight))
    ##IS THIS FABI OR SEISMIC COST??
    #Print results to spreadsheet
    #Unlock model
    SapModel.SetModelIsLocked(False)
    # Delete everything in the model
    ret = SapModel.SelectObj.All(False)
    if ret != 0:
        print('ERROR selecting all')
    ret = SapModel.FrameObj.Delete(Name='', ItemType=2)
    if ret != 0:
        print('ERROR deleting all')
    # Determine total time taken to build current tower
    EndTime = time.time()
    TimeToComputeTower = EndTime - StartTime
    ComputeTimes.append(TimeToComputeTower)
    AverageComputeTime = sum(ComputeTimes) / len(ComputeTimes)
    ElapsedTime = sum(ComputeTimes)
    EstimatedTimeRemaining = (len(AllTowers) - TowerNum) * AverageComputeTime
    if EstimatedTimeRemaining <= 60:
        TimeUnitEstTime = 'seconds'
    elif EstimatedTimeRemaining > 60 and EstimatedTimeRemaining < 3600:
        TimeUnitEstTime = 'minutes'
        EstimatedTimeRemaining = EstimatedTimeRemaining / 60
    else:
        TimeUnitEstTime = 'hours'
        EstimatedTimeRemaining = EstimatedTimeRemaining / 3600

    if ElapsedTime <= 60:
        TimeUnitElaTime = 'seconds'
    elif ElapsedTime > 60 and ElapsedTime < 3600:
        TimeUnitElaTime = 'minutes'
        ElapsedTime = ElapsedTime / 60
    else:
        TimeUnitElaTime = 'hours'
        ElapsedTime = ElapsedTime / 3600
    #Round the times to the nearest 0.1
    AverageComputeTime = int(AverageComputeTime/1) + round(AverageComputeTime - int(AverageComputeTime/1),1)
    EstimatedTimeRemaining = int(EstimatedTimeRemaining/1) + round(EstimatedTimeRemaining - int(EstimatedTimeRemaining/1),1)
    ElapsedTime = int(ElapsedTime/1) + round(ElapsedTime - int(ElapsedTime/1),1)

    # Add FABI to scatter plot
    xdata.append(TowerNum)
    ydata.append(AllFABI[TowerNum-1])
    ScatterPlot.set_xdata(xdata)
    ScatterPlot.set_ydata(ydata)
    plt.xlim(0, TowerNum + 1)
    #plt.ylim(0, max(AllFABI) + max(AllFABI) / 4)
    plt.ylim(0, min(AllFABI) + min(AllFABI) / 4)
    plt.xticks(numpy.arange(min(xdata), max(xdata)+1, 1.0))
    plt.title('Average time per tower: ' + str(AverageComputeTime) + ' seconds\n' + 'Estimated time remaining: ' + str(EstimatedTimeRemaining) + ' ' + TimeUnitEstTime + '\nElapsed time so far: ' + str(ElapsedTime) + ' ' + TimeUnitElaTime)
    plt.draw()
    plt.pause(1e-6)
    plt.show(block=False)
    plt.ion()
    # Increment tower number
    TowerNum += 1

print('\n\nFinished constructing all towers.')

# Write all results to excel spreadsheet
write_to_excel(wb, AllFABI, SaveLoc)
# Close SAP2000
print('Closing SAP2000...')
#SapObject.ApplicationExit(False)
print('FINISHED.')
plt.show(block=True)

